import openpyxl
from Utilites.Login import Login
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
from Applications.Workflows.QBSetup.AppResources import ElementLocators
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

class DC4_Prod_Actions:
    # Check_Capability_Status
    def __init__(self, task_type, lo, username,v_Browser):


        self.v_input_wb = openpyxl.load_workbook(ElementLocators.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name("Input")
        self.v_input_sheet_maps = self.v_input_wb.get_sheet_by_name("Maps")
        self.v_input_sheet_Adaptor_Data_Type_ID = self.v_input_wb.get_sheet_by_name("Adaptor Data Type ID")
        self.log = lo
        self.v_username = username
        self.v_Browser=v_Browser
        self.so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)
        self.login_operations_object = Login(self.v_task_type, self.v_Browser, self.v_input_wb, self.log)
        self.so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)
        self.login_operations_object.login("Launchpad")
        self.login_operations_object.login("DC4 Prod")
        # time.sleep(5)
        # self.v_Browser.get("https://commerce.spscommerce.com/migrator/")
        # time.sleep(10)
        # self.Migrator("960468")
        # time.sleep(10000)

    # def Migrator(self, profile_uid):
    #     # self.v_Browser.get("https://commerce.spscommerce.com/migrator/")
    #     # time.sleep(10)
    #     self.v_Browser.switch_to.frame(0)
    #     time.sleep(5)
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'Search by column')]//select").click()
    #     time.sleep(2)
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'profile_uid')]").click()
    #     # time.sleep(3)
    #     # driver.find_element_by_xpath(".//b").click()
    #     time.sleep(4)
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/a").click()
    #     time.sleep(4)
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").send_keys(
    #         profile_uid)
    #     # driver.find_element_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").click()
    #     time.sleep(3)
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").clear()
    #     self.v_Browser.find_element_by_xpath(
    #         "html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input").send_keys(
    #         profile_uid)
    #     time.sleep(3)
    #     self.v_Browser.find_element_by_xpath(".//ul/li").click()
    #     # self.v_Browser.switch_to.frame(0)
    #     time.sleep(3)
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'Export')]").click()
    #     self.v_Browser.find_element_by_xpath(".//*[contains(text(),'Migrate')]").click()
    #     # self.v_Browser.find_element_by_xpath(".//*[contains(@class,'medium button confirm')]").click()
    #     print("clicked on migrator")


    def Migrator(self, relationship_uid):
        self.v_Browser.get("https://commerce.spscommerce.com/migrator/")
        time.sleep(10)
        self.v_Browser.switch_to.frame(0)
        self.so.click_element_by_xpath(".//*[contains(text(),'Table')]//select")
        self.so.click_element_by_xpath(".//*[contains(text(),'relationship')]")
        self.so.click_element_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/a")
        self.so.send_text_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input",relationship_uid)
        time.sleep(8)
        self.so.send_text_by_xpath("html/body/div[1]/div/div/api-interaction/div/div[4]/label/chosen-select/div/div/div/input",relationship_uid)
        time.sleep(8)
        self.so.click_element_by_xpath(".//ul/li")
        self.so.click_element_by_xpath(".//*[contains(text(),'Export')]")
        time.sleep(15)
        self.so.click_element_by_xpath(".//*[contains(text(),'Migrate')]")
        time.sleep(15)
        print("clicked on migrator")

    def get_maps(self,doc_type,adaptor,retailer_version):
        for i in range(1,self.v_input_sheet_maps.max_row+1):
            doc_type_from_sheet = self.v_input_sheet_maps.cell(row=i, column=1).value
            if str(doc_type_from_sheet)==str(doc_type):
                supplier_version_from_sheet = self.v_input_sheet_maps.cell(row=i, column=2).value
                if adaptor=="Quickbooks" or adaptor=="Fishbowl":
                    supplier_version="7.2"
                if adaptor=="Dwyer" or adaptor=="Peachtree":
                    supplier_version="7"
                if str(supplier_version_from_sheet)==str(supplier_version):
                    retailer_version_from_sheet = self.v_input_sheet_maps.cell(row=i, column=3).value
                    if str(retailer_version_from_sheet)==str(retailer_version):
                        maps=self.v_input_sheet_maps.cell(row=i, column=5).value
                        arr_maps=maps.split(",")
                        return arr_maps

    def get_capability_name(self,adaptor, doc):
        Quickbooks = {"810": ["SPS QuickBooks Adaptor | RSX 7.2 | 810 - Legacy", "106968"],
                      "850": ["SPS QuickBooks Adaptor | RSX 7.2 | 850 - Legacy", "106967"],
                      "875": ["SPS QuickBooks Adaptor | RSX 7.2 | 875 - Legacy", "110240"],
                      "856": ["SPS Quickbooks Adapter RSX 7.2 | OzLink | 856", "135124"]}
        Fishbowl = {"810": ["SPS Fishbowl Adaptor | RSX 7.2 | 810 - Legacy", "109097"],
                    "850": ["SPS Fishbowl Adaptor | RSX 7.2 | 850 - Legacy", "109095"],
                    "875": ["SPS FISHBOWL ADAPTOR | RSX 7.2 | 875 - LEGACY", "112556"],
                    "856": ["SPS Fishbowl Adaptor | RSX 7.2 | 856 - Legacy", "109096"]}
        Dwyer = {"810": ["Dwyer Adaptor V7 810 XML", "31788"], "850": ["Dwyer Adaptor V7 850 XML", "31768"],
                 "875": ["Dwyer Adaptor V7 875 XML", "70230"]}
        Peachtree = {"810": ["Peachtree 810 XML", "73633"], "850": ["Peachtree 850 XML", "73632"],
                     "875": ["Peachtree 875 XML", "78248"]}
        if adaptor == "Quickbooks":
            arr = Quickbooks.get(doc)
            return arr
        if adaptor == "Fishbowl":
            arr = Fishbowl.get(doc)
            return arr
        if adaptor == "Dwyer":
            arr = Dwyer.get(doc)
            return arr
        if adaptor == "Peachtree":
            arr = Peachtree.get(doc)
            return arr

    def Search_By_TPID(self,TPID):

        # self.so.click_element_by_xpath(ElementLocators.Browse_Customers)
        self.v_Browser.get(ElementLocators.DC4_Prod_link)
        self.so.click_element_by_xpath(ElementLocators.Customers_by_TPID)
        self.so.send_text_by_xpath(ElementLocators.TPID_input_box,TPID)
        self.so.click_element_by_xpath(ElementLocators.Search_button)
        Company_name=self.so.get_text_by_xpath(ElementLocators.Company_name)
        Profile_name=self.so.get_text_by_xpath(ElementLocators.Profile_name)
        return Company_name+"$"+Profile_name

    def Open_supplier(self,TPID):
        self.so.click_element_by_xpath(ElementLocators.First_company_name)
        self.so.click_element_by_xpath(ElementLocators.Relationships)
        self.so.click_element_by_xpath(ElementLocators.Relationships_Advanced)
        if self.so.check_exists_by_xpath(ElementLocators.Sender_as_Show_all):
            self.so.click_element_by_xpath(ElementLocators.Sender_as_Show_all)
            self.so.click_element_by_xpath(ElementLocators.Show_all_profile)
            self.so.click_element_by_xpath(ElementLocators.Receiver_as_Show_all)
            self.so.click_element_by_xpath(ElementLocators.Show_all_profile2)
            time.sleep(5)
        Retailer_name=self.so.get_text_by_xpath("//*[contains(text(),'"+str(TPID)+"') and contains(@id,'form1:table3:')]//preceding::span[5]")
        Relationship_UID_as_Sender=self.so.get_text_by_xpath("//*[contains(text(),'"+str(TPID)+"') and contains(@id,'form1:table3:')]//preceding::span[6]")
        Relationship_UID_as_Receiver=self.so.get_text_by_xpath("//*[contains(text(),'"+str(TPID)+"') and contains(@id,'form1:table4:')]//preceding::span[8]")
        self.so.click_element_by_xpath(ElementLocators.Relationship_Overview)
        self.so.send_text_by_xpath(ElementLocators.Search_TP_box,Retailer_name)
        self.so.click_element_by_xpath(ElementLocators.TP_Search_button)
        return Retailer_name+"$"+Relationship_UID_as_Sender+"$"+Relationship_UID_as_Receiver

    def Open_profile(self,Profile_name):
        Company_EDI_Summary=self.so.get_text_by_xpath(ElementLocators.Company_EDI_Summary)
        Trading_Partner_EDI_Summary=self.so.get_text_by_xpath(ElementLocators.Trading_Partner_EDI_Summary)
        self.so.click_element_by_xpath("//*[contains(text(),'"+str(Profile_name)+"')]")
        self.so.click_element_by_xpath(ElementLocators.Show)
        return Company_EDI_Summary+"$"+Trading_Partner_EDI_Summary

    def Add_New_Capability(self,Doc_Type,capability_ID):
        self.so.click_element_by_xpath(ElementLocators.createCapability)
        time.sleep(3)
        all_windows = self.v_Browser.window_handles
        num_of_windows = len(all_windows)
        print(num_of_windows)
        requested_window = all_windows[1]
        self.v_Browser.switch_to.window(requested_window)
        print(self.v_Browser.current_url)
        print("Total Window: " + str(num_of_windows))
        self.v_Browser.switch_to.frame(0)

        if str(Doc_Type)=='850' or str(Doc_Type)=='875':
            service_UID='1007'
        if str(Doc_Type)=='810'or str(Doc_Type)=='855' or str(Doc_Type)== '856':
            service_UID='1006'

        self.so.send_text_by_xpath(ElementLocators.service,service_UID)
        self.so.send_text_by_xpath(ElementLocators.Data_Type,"["+capability_ID+"]")
        self.so.click_element_by_xpath(ElementLocators.submit_create_capability)
        self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])

    def add_existing_capability(self,capability_name,capability_ID,Doc_Type):
        self.so.click_element_by_xpath(ElementLocators.addExistingCapability)
        #switch window
        time.sleep(4)
        all_windows=self.v_Browser.window_handles
        num_of_windows=len(all_windows)
        requested_window=all_windows[1]
        self.v_Browser.switch_to.window(requested_window)

        print(self.v_Browser.current_url)
        print("Total Window: "+str(num_of_windows))

        self.v_Browser.switch_to.frame(0)

        # seq = self.v_Browser.find_elements_by_tag_name('frame')
        # print("seq len")
        # print(len(seq))

        # for index in range(len(seq)):
        #     iframe = self.v_Browser.find_elements_by_tag_name('iframe')[index]
        #     print(iframe)

        # i = 0
        time.sleep(3)
        print("------------------")
        count=self.v_Browser.find_elements_by_xpath("//a[contains(text(),'Show')]")
        print(len(count))
        print("------------------")
        # while True:
        status=''
        for i in range(len(count)):

            required_capability_name = self.so.get_text_by_xpath("//span[@id='form1:table1:" + str(i) + ":outputText6']")
            print(required_capability_name)
            if str(capability_name) == str(required_capability_name):
                print("capability matched...............")
                checkbox = "//input[@id='form1:table1:" + str(i) + ":tableSelectMany1']"
                capability_uid = "//span[@id='form1:table1:" + str(i) + ":outputText2']"
                self.so.click_element_by_xpath(checkbox)
                self.so.click_element_by_xpath(ElementLocators.cap_choose_btn)
                # self.v_Browser.close()
                self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])
                status='capability selected'
                break
            else:
                status='Capability not available'
        if status=="Capability not available":
            self.v_Browser.close()
            self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])
            self.Add_New_Capability(Doc_Type,capability_ID)
            return "filebroker_required"
        else:
            return "filebroker_not_required"
        print(status)


            # print("in while loop")
            # flag = self.so.check_exists_by_xpath(required_capability_name)
            # val=self.so.get_text_by_xpath(required_capability_name)
            # print(val)
            # if self.so.check_exists_by_xpath(required_capability_name):
            #     print("in True loop")
            #     name_from_UI = self.so.get_text_by_xpath(required_capability_name)
            #     if name_from_UI == capability_name:
            #         print("capability matched...............")
            #         checkbox="//input[@id='form1:table1:"+str(i)+":tableSelectMany1']"
            #         capability_uid="//span[@id='form1:table1:"+str(i)+":outputText2']"
            #         self.so.click_element_by_xpath(checkbox)
            #         # self.so.click_element_by_xpath(ElementLocators.choose_btn)
            #         time.sleep(3)
            #         # self.v_Browser.close()
            #         print("about to close")
            #         time.sleep(1190)
            #         break
            #
            # if self.so.check_exists_by_xpath(required_capability_name)==False:
            #     print("in false loop")
            #     time.sleep(3)
            #     self.v_Browser.close()
            #     # requested_window = all_windows[0]
            #     # self.v_Browser.switch_to.window(requested_window)
            #     # self.so.click_element_by_xpath(ElementLocators.Relationships)
            #     time.sleep(1900)
                # capability not available
                # click on
                # break
            # i += 1

        # if flag==False:
        #     break

    def Toggle_profile_capability(self,capability_name):
        self.so.click_element_by_xpath(ElementLocators.Profiles_Tab)
        time.sleep(3)
        i = 0
        action = ''
        while True:
            Datatype_Name = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText14']"
            flag = self.so.check_exists_by_xpath(Datatype_Name)
            if flag == True:
                name_from_UI = self.so.get_text_by_xpath(Datatype_Name)
                if name_from_UI == capability_name:
                    status = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText13']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']	"
                    Extentions = "//a[@id='form1:table1:0:table2:" + str(i) + ":extensionPopup']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']	"
                    Status = self.so.get_text_by_xpath(status)
                    self.so.click_element_by_xpath(checkbox)
                    self.so.click_element_by_xpath(ElementLocators.Toggle_profile_capability_status)
                    self.so.click_element_by_xpath(ElementLocators.Profiles_Tab)

                    # if Status == "Active":
                    #     action = "do_nothing"
                    #     break
                    # if Status == "Disabled":
                    #     action = "do_nothing"
                        # check exxtentions
                        # self.so.click_element_by_xpath(checkbox)
                    break
            if flag == False:
                # capability not available
                # click on
                # self.add_existing_capability(capability_name, capability_ID, Doc_Type)

                break
            i += 1

    def filebroker(self,capability_name,doc_type,ftp_name):
    # def filebroker(self):

        # added_cap_UID="2820910"
        # doc_type="850"
        # ftp_name="Demo"


        self.so.click_element_by_xpath(ElementLocators.Capabilities)
        # self.so.send_text_by_xpath(ElementLocators.Capability_UID_textbox,added_cap_UID)
        self.so.send_text_by_xpath(ElementLocators.Datatype_Name_textbox,capability_name)
        self.so.click_element_by_xpath(ElementLocators.search_capability)
        self.so.click_element_by_xpath(ElementLocators.editFilebrokerSettingsButton)
        time.sleep(5)
        self.v_Browser.switch_to.window(self.v_Browser.window_handles[1])
        self.v_Browser.switch_to.frame(0)
        self.so.click_element_by_xpath(ElementLocators.add_filebroker_btn)
        time.sleep(4)
        if doc_type=='850' or doc_type=='875':
            Archive_Directory_path="/u01/Arc/ftp/vendor/"+ftp_name+"/out"
            Directory_path="/u01/ftp/vendor/"+ftp_name+"/out"
            FilenameMacro="PO%p"

            time.sleep(2)
            select = Select(self.v_Browser.find_element_by_id('_idJsp4:_idJsp8'))
            select.select_by_visible_text('DC4 - Read or write files from/to DC4')

            time.sleep(2)
            select = Select(self.v_Browser.find_element_by_id('_idJsp4:_idJsp39'))
            select.select_by_visible_text('File - Read or write files from/to internal comms server')
            time.sleep(4)
            #
            # self.so.click_element_by_xpath(ElementLocators.Source_Channel)
            # self.so.click_element_by_xpath(ElementLocators.Source_Channel_text)
            # self.so.click_element_by_xpath("//*[contains(text(),'Destination Channel')]")
            # time.sleep(3)
            # self.so.click_element_by_xpath(ElementLocators.Destination_Channel)
            # time.sleep(3)
            # self.so.click_element_by_xpath(ElementLocators.Destination_Channel_text)
            # self.so.click_element_by_xpath("//*[contains(text(),'Destination Channel')]")
            # time.sleep(2)

            # self.so.send_text_by_xpath(ElementLocators.Source_Channel,"DC4")
            # self.so.click_element_by_xpath(ElementLocators.Destination_Channel)
            # self.so.send_text_by_xpath(ElementLocators.Destination_Channel,"File - Read or write files from/to internal comms server")
            self.so.click_element_by_xpath(ElementLocators.continue_btn)
            time.sleep(4)
            self.so.send_text_by_xpath(ElementLocators.ArchiveDirectory_fito,Archive_Directory_path)
            self.so.send_text_by_xpath(ElementLocators.Directory_fito,Directory_path)
            self.so.send_text_by_xpath(ElementLocators.FilenameMacro_fito,FilenameMacro)
            self.so.click_element_by_xpath(ElementLocators.Save_Changes_btn)
            time.sleep(3)
            self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])


        if doc_type=='810' or doc_type=='856':
            Archive_Directory_path="/u01/Arc/ftp/vendor/"+ftp_name+"/in"
            Directory_path="/u01/ftp/vendor/"+ftp_name+"/in"
            status="Active"
            if doc_type=='810':
                FileSpecification="IN*"
            if doc_type=='856':
                FileSpecification="ASN*"

            time.sleep(2)
            select = Select(self.v_Browser.find_element_by_id('_idJsp4:_idJsp8'))
            select.select_by_visible_text('File - Read or write files from/to internal comms server')
            time.sleep(2)
            select = Select(self.v_Browser.find_element_by_id('_idJsp4:_idJsp39'))
            select.select_by_visible_text('DC4 - Read or write files from/to DC4')
            time.sleep(2)
            select = Select(self.v_Browser.find_element_by_id('_idJsp4:_idJsp70'))
            select.select_by_visible_text('Active')
            time.sleep(2)
            self.so.click_element_by_xpath(ElementLocators.continue_btn)



            # self.so.click_element_by_xpath(ElementLocators.Source_Channel)
            # self.so.click_element_by_xpath((ElementLocators.Destination_Channel_text))
            # self.so.click_element_by_xpath(ElementLocators.Destination_Channel)
            # self.so.click_element_by_xpath(ElementLocators.Source_Channel_text)
            #
            # self.so.send_text_by_xpath(ElementLocators.Source_Channel,"File - Read or write files from/to internal comms server")
            # self.so.send_text_by_xpath(ElementLocators.Destination_Channel,"DC4 - Read or write files from/to DC4")
            # self.so.click_element_by_xpath(ElementLocators.Status)
            # self.so.click_element_by_xpath(ElementLocators.Status_text)


            # self.so.send_text_by_xpath(ElementLocators.Status,status)

            time.sleep(3)
            self.so.send_text_by_xpath(ElementLocators.ArchiveDirectory_fifrom,Archive_Directory_path)
            self.so.send_text_by_xpath(ElementLocators.Directory_fifrom,Directory_path)
            self.so.send_text_by_xpath(ElementLocators.FileSpecification,FileSpecification)
            self.so.click_element_by_xpath(ElementLocators.Save_Changes_btn)
            time.sleep(3)
            self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])

    def Add_Extentions(self, capability_name,arr_maps,Doc_Type,ftp_name):
        total_extentions = len(arr_maps)
        time.sleep(5)
        self.so.click_element_by_xpath(ElementLocators.Profiles_Tab)
        time.sleep(3)
        i = 0
        action = ''
        while True:
            Datatype_Name = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText14']"
            flag = self.so.check_exists_by_xpath(Datatype_Name)
            if flag == True:
                name_from_UI = self.so.get_text_by_xpath(Datatype_Name)
                if name_from_UI == capability_name:
                    status = "//span[@id='form1:table1:0:table2:" + str(i) + ":outputText13']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']"
                    Extentions = "//a[@id='form1:table1:0:table2:" + str(i) + ":extensionPopup']"
                    checkbox = "//input[@id='form1:table1:0:table2:" + str(i) + ":tableSelectMany1']"
                    Status = self.so.get_text_by_xpath(status)
                    added_cap_UID=self.so.get_text_by_xpath("//span[@id='form1:table1:0:table2:" + str(i) + ":outputText24']")

                    self.so.click_element_by_xpath(Extentions)
                    time.sleep(4)
                    self.v_Browser.switch_to.window(self.v_Browser.window_handles[1])
                    self.v_Browser.switch_to.frame(0)

                    for i in range(total_extentions):

                        self.so.click_element_by_xpath(ElementLocators.Add_Extention_btn)

                        self.so.click_element_by_xpath(ElementLocators.map_1080)
                        self.so.click_element_by_xpath(ElementLocators.map_1080_choose)

                    for i in range(total_extentions):
                        self.so.click_element_by_xpath("//a[@id='form1:table1dd"+str(i)+"']//img")

                    for i in range(total_extentions):
                        self.so.send_text_by_xpath("//input[@id='form1:table1:"+str(i)+":table2:0:outputText22']",arr_maps[i])

                    self.so.click_element_by_xpath(ElementLocators.Extention_Save_Changes)
                    time.sleep(3)
                    self.v_Browser.switch_to.window(self.v_Browser.window_handles[0])

                    ############################################################### self.filebroker(added_cap_UID, Doc_Type, ftp_name)



                    # if Status == "Active":
                    #     action = "do_nothing"
                    #     break
                    # if Status == "Disabled":
                    #     action = "do_nothing"
                    # check exxtentions
                    # self.so.click_element_by_xpath(checkbox)
                    break
            if flag == False:
                # capability not available
                # click on
                # self.add_existing_capability(capability_name, capability_ID, Doc_Type)

                break
            i += 1
        return added_cap_UID

    def Check_Capability_Status(self,capability_name,capability_ID,Doc_Type,arr_maps,ftp_name):
        time.sleep(4)
        i=0
        action=''
        while True:
            Datatype_Name="//span[@id='form1:table1:0:table2:"+str(i)+":outputText14']"
            flag=self.so.check_exists_by_xpath(Datatype_Name)
            if flag==True:
                name_from_UI=self.so.get_text_by_xpath(Datatype_Name)
                if name_from_UI==capability_name:
                    status="//span[@id='form1:table1:0:table2:"+str(i)+":outputText13']"
                    checkbox="//input[@id='form1:table1:0:table2:"+str(i)+":tableSelectMany1']	"
                    Extentions="//a[@id='form1:table1:0:table2:"+str(i)+":extensionPopup']"
                    checkbox="//input[@id='form1:table1:0:table2:"+str(i)+":tableSelectMany1']	"
                    Status=self.so.get_text_by_xpath(status)
                    print(name_from_UI)
                    print(Status)

                    if Status=="Active":
                        action="do_nothing"
                        break
                    if Status=="Disabled":
                        action = "do_nothing"
                        #check exxtentions
                        # self.so.click_element_by_xpath(checkbox)
                    break
            if flag==False:
                #capability not available
                #click on
                filebroker_required_or_not=self.add_existing_capability(capability_name,capability_ID,Doc_Type)
                self.Toggle_profile_capability(capability_name)
                self.Add_Extentions(capability_name,arr_maps,Doc_Type,ftp_name)
                if filebroker_required_or_not=="filebroker_required":
                    self.filebroker(capability_name, Doc_Type, ftp_name)
                break
            i += 1


