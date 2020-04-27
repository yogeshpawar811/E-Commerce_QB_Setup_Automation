import openpyxl
from Applications.Workflows.QBSetup.AppResources import ElementLocators
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
from Applications.Workflows.ServiceTimeCard.TaskCategorization.Script.Salesforce import Salesforce
import time
from Applications.Workflows.QBSetup.Script.DC4_Prod_Actions import DC4_Prod_Actions

#filebroker only available for = 850,875,810,856


class QB_Setup:

    def __init__(self, task_type, lo, username):



        
        print("in init")

        self.v_input_wb = openpyxl.load_workbook(ElementLocators.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name("Input")
        self.v_input_sheet_maps = self.v_input_wb.get_sheet_by_name("Maps")
        self.log = lo
        self.v_username = username
        self.v_Browser = webdriver.Chrome(ElementLocators.BROWSER_DRIVER)
        self.v_Browser.maximize_window()
        self.dc4_prod=DC4_Prod_Actions(task_type, lo, username,self.v_Browser)


    def execute_main(self):

        for i in range(2,self.v_input_sheet.max_row+1):
            TPID=self.v_input_sheet.cell(row=i,column=1).value
            Doc_Type=self.v_input_sheet.cell(row=i,column=2).value
            Adaptor=self.v_input_sheet.cell(row=i,column=3).value
            Retailer_version=self.v_input_sheet.cell(row=i,column=4).value
            ftp_name=self.v_input_sheet.cell(row=i,column=5).value
            arr_maps = self.dc4_prod.get_maps(Doc_Type,Adaptor,Retailer_version)
            print(arr_maps)

            if arr_maps!= None:
                cap_and_ID=self.dc4_prod.get_capability_name(Adaptor,str(Doc_Type))
                print(cap_and_ID)
                capability_name=cap_and_ID[0]
                # capability_name="TARGET Corporation 875 FEDS RSX1a"
                capability_ID=cap_and_ID[1]
                Company_name_and_Profile_name=self.dc4_prod.Search_By_TPID(TPID)
                Supplier_name=Company_name_and_Profile_name.split("$")[0]
                print("Supplier_name:"+Supplier_name)
                Profile_name=Company_name_and_Profile_name.split("$")[1]
                print("Profile_name:"+Profile_name)
                Retailer_name_Relationship_UID_as_Sender_Relationship_UID_as_Receiver=self.dc4_prod.Open_supplier(TPID)

                Retailer_name=Retailer_name_Relationship_UID_as_Sender_Relationship_UID_as_Receiver.split("$")[0]
                print("Retailer_name:"+Retailer_name)
                Relationship_UID_as_Sender=Retailer_name_Relationship_UID_as_Sender_Relationship_UID_as_Receiver.split("$")[1]
                print("Relationship_UID_as_Sender:"+str(Relationship_UID_as_Sender))
                Relationship_UID_as_Receiver=Retailer_name_Relationship_UID_as_Sender_Relationship_UID_as_Receiver.split("$")[2]
                print("Relationship_UID_as_Receiver:"+str(Relationship_UID_as_Receiver))
                Company_EDI_Summary_Trading_Partner_EDI_Summary=self.dc4_prod.Open_profile(Profile_name)
                Company_EDI_Summary=Company_EDI_Summary_Trading_Partner_EDI_Summary.split("$")[0]
                print("Company_EDI_Summary:"+Company_EDI_Summary)
                Trading_Partner_EDI_Summary=Company_EDI_Summary_Trading_Partner_EDI_Summary.split("$")[1]
                print("Trading_Partner_EDI_Summary:"+Trading_Partner_EDI_Summary)
                self.dc4_prod.Check_Capability_Status(capability_name,capability_ID,Doc_Type,arr_maps,ftp_name)
                self.dc4_prod.Migrator(Relationship_UID_as_Sender)
                self.dc4_prod.Migrator(Relationship_UID_as_Receiver)
            elif arr_maps== None:
                print("Please add valid maps for this connection in excel sheet !!!")