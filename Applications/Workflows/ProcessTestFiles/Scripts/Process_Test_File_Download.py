import time
import math
import datetime
import openpyxl
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.ProductionDataMonitoring.AppResources import LocalElementLocator
from Utilites.TransactionTrackerOperations_Old import TransactionTrackerOperations
from Utilites.TransactionTrackerOperations_Old import TransactionTrackerOperations
from Utilites.ExcelOperations import ExcelOperations
#from Applications.Workflows.ProcessTestFiles.AppResources import CommonLocators
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.ProductionDataMonitoring.AppResources import LocalElementLocator
from Utilites.TransactionTrackerOperations_Old import TransactionTrackerOperations
# from Applications.Workflows.ProcessTestFile.Scripts.ProcessTestFile_Utilities import ProcessTestFile_Utilities
from Utilites.TransactionTrackerOperations_Old import TransactionTrackerOperations
from Utilites.ExcelOperations import ExcelOperations
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from Applications.Workflows.ProcessTestFiles.Scripts.Process_Test_Files_Utility import Process_Test_Files_Utility
import shutil


class TT_File_Download:

    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        #chrome_options = webdriver.ChromeOptions()
        #chrome_options.add_argument("--incognito")
        #self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER, chrome_options=chrome_options)

        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.lo = lo
        self.v_username = username
        # self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name("PROCESS_TEST_FILES_INPUT")
        #self.lo = lo
        #self.v_username = username


    def process(self, path):
        #tto = TransactionTrackerOperations
        print("in task filter method")
        #rf = ReportFileUtility(self.v_task_type)
        # self.v_input_sheet = self.v_input_wb.get_sheet_names("InputData")
        print("input data sheet selected")
        row_count = self.v_input_sheet.max_row
        print(row_count)
        eo = ExcelOperations(self.v_task_type, self.v_input_sheet)
        supplier_name = self.v_input_sheet.cell(row=2, column=1).value
        retailer_name = self.v_input_sheet.cell(row=2, column=2).value
        doc_type = self.v_input_sheet.cell(row=2, column=3).value
        date = self.v_input_sheet.cell(row=2, column=4).value
        ptfu = Process_Test_Files_Utility(self.v_task_type, self.v_driver, self.lo)
        so = SeleniumOperations(self.v_task_type,self.v_driver,self.lo)
        time.sleep(3)
        self.v_driver.switch_to.frame(0)
        ptfu.search_by_names(supplier_name,retailer_name,doc_type,date)
        # ptfu.search_by_names("AURORA ORGANIC","Topco","850","09/20/2018")
        ptfu.get_five_parcels()

        path = 'D:\processtestfiles.txt'

        with open(path) as f:
            for line in f:
                parcel=line.replace('\n', '')
                print("Download Parcel ID : "+str(parcel))
                #self.v_driver.switch_to.frame(0)
                ptfu.search_by_parcel_id(str(parcel))

                #generated_parcel_id = ptfu.save_page_source_and_generate_parcel_id_from_dom()

                #xpath_for_click = ptfu.generate_xpath_for_TT(generated_parcel_id)

                so.click_element_by_xpath("//aside[text()='Transformations']/following::a[@title='View'][1]/div/div")
                # time.sleep(1)
                # downloaded_parcel=so.get_text_by_xpath("//aside[text()='Transformations']/following::a[@title='View'][1]/div/div/following::span[2]")
                # downloaded_parcel_path="C:\\Users\\yogesh.pawar\\Downloads\\"+downloaded_parcel+".dat"
                # copy_to_path="D:\\ProcessTestFiles"
                element = WebDriverWait(self.v_driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "html/body/div[1]/section/section/div/div/section/div/div[3]/div[1]/div/div[1]/div/div/div/a/i")))
                element.click()
                time.sleep(2)
                # shutil.copy(downloaded_parcel_path, copy_to_path)

                if 'str' in line:
                    #time.sleep(4)
                    break
    #
    # def get_five_parcels(self):
    #     print("in download file method")
    #     time.sleep(4)
    #     # driver.switch_to.frame(0)
    #     parcels = ''
    #     for i in range(5):
    #         parcel_id = self.v_driver.find_element_by_xpath(
    #             ".//*[@id='parentTablesContainer']/div[1]/table/tbody/tr[" + str(i + 1) + "]/td[2]/span").text
    #         parcels = parcels + parcel_id + "\n"
    #     print(parcels)
    #     file = open("D://processtestfiles.txt", "w+")
    #     file.write(parcels)


    def execute_main(self):
        # self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        # self.v_driver.maximize_window()
        self.lo.log_to_file("INFO", "Login in to Launchpad")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        time.sleep(5)
        print("Browser opened,login,open TT")
        self.process(self.v_input_wb)
        print("Task filter completed")
        #self.get_five_parcels()
        #self.v_driver.close()
        #time.sleep(60)
        #self.TaskFilter(self.v_input_wb, 'InProgress')
        # self.v_driver.close()