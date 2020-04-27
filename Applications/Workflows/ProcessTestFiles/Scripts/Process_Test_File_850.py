'''

@ Author - Aditya Datar
@ Creation date - 09/14/2018
@ Description - Main Script of Review Setup
'''
from Applications.Workflows.ProcessTestFiles.Scripts.Process_Test_Files_Utility import Process_Test_Files_Utility
from Utilites import AppConstants
from selenium import webdriver
from Utilites import Login
from Utilites.SeleniumOperations import SeleniumOperations
import time
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from Utilites.ExcelOperations import ExcelOperations
from Utilites.TransactionTrackerOperations_Old import TransactionTrackerOperations
import openpyxl


class Process_Test_file_850:

    def __init__(self, task_type, lo, username,input_wb, v_driver):
        self.v_task_type = task_type
        self.v_driver = v_driver
        self.v_input_wb = input_wb
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')
        # self.v_data_wb = openpyxl.load_workbook(AppConstants.PROCESS_TEST_FILES_INPUT_PATH)
        self.lo = lo
        self.v_username = username

    # Method is used to download the PO files from TT prod
    def process(self, path):
        tto = TransactionTrackerOperations
        print("in task filter method")
        #rf = ReportFileUtility(self.v_task_type)
        # self.v_input_sheet = self.v_input_wb.get_sheet_names("InputData")
        print("input data sheet selected")
        row_count = self.v_data_sheet.max_row
        print(row_count)

        eo = ExcelOperations(self.v_task_type, self.v_data_sheet)
        supplier_name = self.v_data_sheet.cell(row=2, column=1).value
        retailer_name = self.v_data_sheet.cell(row=2, column=2).value
        doc_type = self.v_data_sheet.cell(row=2, column=3).value
        date = self.v_data_sheet.cell(row=2, column=4).value
        ptfu = Process_Test_Files_Utility(self.v_task_type, self.lo, self.v_driver,self.v_input_wb)
        so = SeleniumOperations(self.v_task_type,self.v_driver,self.lo)
        time.sleep(3)
        self.v_driver.switch_to.frame(0)
        time.sleep(2)
        ptfu.search_by_names(supplier_name,retailer_name,doc_type,date)
        # ptfu.search_by_names("CAULIPOWER","Amazon.com","850","09/20/2018")
        ptfu.get_five_parcels()

        path = '../Applications/Workflows/ProcessTestFiles/AppResources/parcelIDsforSearch.txt'
        #path = 'D:\parcelIDsforSearch.txt'


        with open(path) as f:
            parcel_id_for_vali_and_autoit=''
            for line in f:
                parcel=line.replace('\n', '')
                print("Download Parcel ID : "+str(parcel))
                #self.v_driver.switch_to.frame(0)
                ptfu.search_by_parcel_id(str(parcel))

                #generated_parcel_id = ptfu.save_page_source_and_generate_parcel_id_from_dom()

                #xpath_for_click = ptfu.generate_xpath_for_TT(generated_parcel_id)
                parcel_id_for_validation_and_autoit = so.get_text_by_xpath(
                    "//aside[text()='Transformations']/following::a[@title='View']/div[1]/following::span[2]")
                print("downloaded parcel ID :" + str(parcel_id_for_validation_and_autoit))
                so.click_element_by_xpath("//aside[text()='Transformations']/following::a[@title='View'][1]/div/div")

                parcel_id_for_vali_and_autoit = parcel_id_for_vali_and_autoit + parcel_id_for_validation_and_autoit + "\n"
                print(parcel_id_for_validation_and_autoit)




                element = WebDriverWait(self.v_driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "html/body/div[1]/section/section/div/div/section/div/div[3]/div[1]/div/div[1]/div/div/div/a/i")))
                element.click()
                time.sleep(2)
                # shutil.copy(downloaded_parcel_path, copy_to_path)

                if 'str' in line:
                    #time.sleep(4)
                    break


            file = open("../Applications/Workflows/ProcessTestFiles/AppResources/parcel_id_for_validation_and_autoit.txt", "w+")
            file.write(parcel_id_for_vali_and_autoit)


    def execute_main(self,v_supplier,v_retailer, v_doc_type, row_count):

        service_name = "FItoService"
        ptf = Process_Test_Files_Utility(self.v_task_type,self.lo, self.v_driver, self.v_input_wb)
        #ptf.supplier_setup_check(v_supplier,v_retailer,v_doc_type,row_count, service_name)
        #ptf.retailer_setup_check(v_supplier,v_retailer,v_doc_type,row_count)
        # self.v_driver.switch_to.window(self.v_driver.window_handles[0])
        # time.sleep(5)
        print("Browser opened,login,open TT")
        self.process(self.v_input_wb)
        time.sleep(2)


        # Back to DC4 Page
        #self.v_driver.switch_to.window(self.v_driver.window_handles[1])