'''

@ Author - Aditya Datar
@ Creation date - 09/14/2018
@ Description - Main Script of Review Setup
'''

import openpyxl
from easygui import *
from builtins import set
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.ExcelOperations import ExcelOperations
from Utilites.SeleniumOperations import SeleniumOperations
from selenium.common.exceptions import NoSuchElementException
from Applications.Workflows.ProcessTestFiles.AppResources import CommonLocators


import time
import re


class Process_Test_Files_Utility:
    def __init__(self, task_type, lo, driver, input_wb):
        self.v_task_type = task_type
        self.v_driver = driver
        self.v_input_wb = input_wb
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')
        # self.v_data_wb = openpyxl.load_workbook(AppConstants.PROCESS_TEST_FILES_INPUT_PATH)
        self.lo = lo


    # Method to check the supplier profile and review the setup to check whether the 850 FI capability is present or not
    def supplier_setup_check(self, v_supplier, v_retailer, v_document_type, current_row, service_name):

        v_document_type = str(v_document_type)
        self.lo.log_to_file("INFO", "Login in to DC4 Pre_Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)

        # lg.login("DC4 PreProd")
        #
        # # Switch to other window
        # self.v_driver.execute_script("window.open('about:blank', 'tab2');")
        # self.v_driver.switch_to.window("tab2")
        # time.sleep(2)
        #
        # lg.login("Launchpad")
        # time.sleep(5)
        #
        # self.v_driver.maximize_window()

        eo = ExcelOperations(self.v_task_type, self.v_data_sheet)

        self.lo.log_to_file("INFO", "Login in to DC4 Pre_Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)

        so.click_element_by_xpath(AppConstants.DC4_TAB)

        dc = DC4_Utility(self.v_task_type, self.v_driver, self.lo)
        so.send_text_by_xpath(AppConstants.DC4_COMPANY_NAME_TEXT_FIELD,v_supplier)
        so.click_element_by_xpath(AppConstants.DC4_COMPANY_NAME_SEARCH_CLICK)
        select_supplier = so.click_element_by_xpath(CommonLocators.Company_Select_Link)

        # Check the retationship
        click_relationship = self.v_driver.find_element_by_xpath(CommonLocators.Relationship_Tab).click()
        search_retailer_relationship = self.v_driver.find_element_by_xpath(
            CommonLocators.Trading_Partner_Name_text_Field).send_keys(v_retailer)

        self.v_driver.find_element_by_xpath(CommonLocators.click_search).click()


        # retailer_profile_name = self.v_driver.find_element_by_xpath(CommonLocators.Trading_Partner_Profile_Name).text
        # print(retailer_profile_name)

        supplier_EDI_Data = self.v_driver.find_element_by_xpath(CommonLocators.Supplier_EDI_Info).text
        retailer_EDI_Data = self.v_driver.find_element_by_xpath(CommonLocators.Retailer_EDI_Info).text

        # Supplier EDI Info
        supplier_EDI_Data_list_1 = supplier_EDI_Data.split('GRP_ID:')
        supplier_EDI_Data_list_2 = supplier_EDI_Data.split('Qual:')

        supplier_qaul_ID = supplier_EDI_Data_list_2[1].split('ID:')[0]
        supplier_qaul_ID = re.sub('\W+', '', supplier_qaul_ID)
        eo.set_value(current_row, 5, supplier_qaul_ID)
        self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)

        supplier_ISA_ID = supplier_EDI_Data_list_1[0].split('ID:')[2]
        supplier_ISA_ID = re.sub('\W+', '', supplier_ISA_ID)
        eo.set_value(current_row, 6, supplier_ISA_ID)
        self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)

        supplier_Grp_ID = supplier_EDI_Data_list_1[1]
        supplier_Grp_ID = re.sub('\W+', '', supplier_Grp_ID)

        # Retailer EDI Info
        retailer_EDI_Data_list_1 = retailer_EDI_Data.split('GRP_ID:')
        retailer_EDI_Data_list_2 = retailer_EDI_Data.split('Qual:')

        retailer_qaul_ID = retailer_EDI_Data_list_2[1].split('ID:')[0]
        retailer_qaul_ID = re.sub('\W+', '', retailer_qaul_ID)
        eo.set_value(current_row, 7, retailer_qaul_ID)
        self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)

        retailer_ISA_ID = retailer_EDI_Data_list_1[0].split('ID:')[2]
        retailer_ISA_ID = re.sub('\W+', '', retailer_ISA_ID)
        eo.set_value(current_row, 8, retailer_ISA_ID)
        self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)

        retailer_Grp_ID = retailer_EDI_Data_list_1[1]
        retailer_Grp_ID = re.sub('\W+', '', retailer_Grp_ID)

        v_trading_partner_name = self.v_driver.find_element_by_xpath(CommonLocators.Trading_Partner_Name).text
        print(v_trading_partner_name)

        if v_trading_partner_name == v_retailer:

            # Select the Supplier company profile and check for the FI setup
            select_supplier_profile = self.v_driver.find_element_by_xpath(CommonLocators.Supplier_Profile_Link).click()

        elif v_trading_partner_name == "No trading partners found.":

            print("No Retailer Found")
            return


        # Profile Page Display
        show_option_click = self.v_driver.find_element_by_xpath(CommonLocators.show_click).click()

        counter = 2
        path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr['+str(counter)+']/td[10]'
        while(so.check_exists_by_xpath(path)):

            service = so.get_text_by_xpath('//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr['+str(counter)+']/td[10]')
            doc_tpe = so.get_text_by_xpath('//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr['+str(counter)+']/td[3]')

            if service == service_name:
                if doc_tpe == v_document_type:


                    eo.set_value(current_row, 11, "FI Capability is present")

                    supplier_version = so.get_text_by_xpath('//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr['+str(counter)+']/td[7]')
                    eo.set_value(current_row,9,supplier_version)
                    self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)

                    click_extension = so.click_element_by_xpath(
                        '//*[@id="form1:table1:0:table2:' + str(counter) + ':extensionPopup"]')

                    # Process_Test_Files_Utility.add_extensions(counter)
                    break
                else:
                    eo.set_value(current_row, 11, "FI is not present")

            self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
            counter = counter + 1
            path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[10]'

        time.sleep(15)

    # Method to check the retailer end to verify that 850 capability is present or not
    def retailer_setup_check(self, v_supplier, v_retailer, v_document_type, current_row):
        v_document_type_1 = str(v_document_type)
        v_document_type_2 = '875'
        self.lo.log_to_file("INFO", "Login in to DC4 Pre_Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)

        eo = ExcelOperations(self.v_task_type, self.v_data_sheet)

        self.lo.log_to_file("INFO", "Login in to DC4 Pre_Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)

        so.click_element_by_xpath(AppConstants.DC4_TAB)

        dc = DC4_Utility(self.v_task_type, self.v_driver, self.lo)
        self.v_driver.find_element_by_xpath(AppConstants.DC4_COMPANY_NAME_TEXT_FIELD).send_keys(v_retailer)
        self.v_driver.find_element_by_xpath(AppConstants.DC4_COMPANY_NAME_SEARCH_CLICK).click()
        select_retailer = self.v_driver.find_element_by_xpath(CommonLocators.Company_Select_Link).click()

        # Check the retationship
        click_relationship = so.click_element_by_xpath(CommonLocators.Relationship_Tab)
        search_retailer_relationship = so.send_text_by_xpath(
            CommonLocators.Trading_Partner_Name_text_Field,v_supplier)
        so.click_element_by_xpath(CommonLocators.click_search)

        # Select the Retailer company profile and check for the FI setup
        select_retailer_profile = so.click_element_by_xpath(CommonLocators.Supplier_Profile_Link)

        # Profile Page Display
        show_option_click = so.click_element_by_xpath(CommonLocators.show_click)

        counter = 2
        path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[10]'
        while (so.check_exists_by_xpath(path)):

            service = so.get_text_by_xpath(
              '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[10]')
            doc_tpe = so.get_text_by_xpath(
                '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[3]')
            print(doc_tpe)



            if doc_tpe == v_document_type_1:
                if service != 'DoNotRoute':
                    print("850 capability is available")

                    retailer_version = so.get_text_by_xpath(
                        '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[8]')
                    eo.set_value(current_row, 10, retailer_version)
                    self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
                    break

                elif service == 'DoNotRoute':
                    print("850 capability is not available")

            elif doc_tpe == v_document_type_2:
                if service != 'DoNotRoute':
                    print("875 capability is present")

                    retailer_version = so.get_text_by_xpath(
                        '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[8]')
                    eo.set_value(current_row, 10, retailer_version)
                    self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
                    break

                elif service == 'DoNotRoute':
                    print("875 capability is present")

            counter = counter + 1
            path = '//*[@id="form1:table1:0:table2"]/table/tbody/tr[3]/td/table/tbody/tr[' + str(counter) + ']/td[10]'


        time.sleep(10)

        # self.v_driver.close()



    def select_customer(self, customer_type, customer_name):
        #self.driver.switch_to.frame(0)
        self.lo.log_to_file("INFO", "Executing method 'select_customer' from TransactionTrackerOperations")
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        print("Search for: " + str(customer_name))

        time.sleep(2)
        # driver.switch_to.frame(0)
        if customer_type == "Company":
            # self.driver.switch_to.frame(0)
            so.send_text_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[1]/div/div[1]/label/chosen-company/div/ul/li/input",customer_name)

        if customer_type == "Trading Partner":
            # self.driver.switch_to.frame(0)
            # TP_name = self.driver.find_element_by_xpath(
            #     "html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[1]/div/div[4]/label/chosen-trading-partner/div/ul/li/input")
            # time.sleep(2)
            print("================================"+customer_name)
            so.send_text_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[1]/div/div[4]/label/chosen-trading-partner/div/ul/li/input",customer_name)
            # driver.find_element_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[1]/div/div[4]/label/chosen-trading-partner/div/ul").click()
           #TP_name.send_keys(customer_name)

        time.sleep(3)
        count = self.v_driver.find_elements_by_xpath(".//*[contains(@id,'ui-select-choices-row-')]")
        for i in range(len(count)):
            if customer_type == "Company":
                customer_from_TT = self.v_driver.find_element_by_xpath(
                    ".//*[@id='ui-select-choices-row-0-" + str(i) + "']").text
            if customer_type == "Trading Partner":
                customer_from_TT = self.v_driver.find_element_by_xpath(
                    ".//*[@id='ui-select-choices-row-1-" + str(i) + "']").text
            if customer_name == customer_from_TT:
                if customer_type == "Company":
                    so.click_element_by_xpath(".//*[@id='ui-select-choices-row-0-" + str(i) + "']")
                    #self.driver.find_element_by_xpath(".//*[@id='ui-select-choices-row-0-" + str(i) + "']").click()
                if customer_type == "Trading Partner":
                    so.click_element_by_xpath(".//*[@id='ui-select-choices-row-1-" + str(i) + "']")
                    #self.driver.find_element_by_xpath(".//*[@id='ui-select-choices-row-1-" + str(i) + "']").click()
                print("Found matching customer name at position: " + str(i + 1))
        time.sleep(2)

    def search_by_names(self,supplier,retailer,doc_type,date):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)

        print("in search_by_names method")
        self.select_customer("Company",supplier)
        time.sleep(3)
        self.select_customer("Trading Partner",retailer)
        time.sleep(2)

        so.send_text_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[2]/div/div[2]/div/div/div[1]/label/input",date)

        time.sleep(2)

        time.sleep(2)
        service=self.v_driver.find_element_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[2]/div/div[1]/label/select")
        #so.send_text_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[2]/div/div[1]/label/select", "DC4Router")
        service.send_keys("DC4Router")

        so.send_text_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[2]/div/div[2]/div/div[3]/label/input", doc_type)


        search_btn = self.v_driver.find_element_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[5]/div/button")
        search_btn.click()
        #time.sleep(3)
        clear_btn = self.v_driver.find_element_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[5]/div/a")
        #clear_btn.click()

    def get_five_parcels(self):
        print("in get_five_parcels method")
        time.sleep(4)
        # driver.switch_to.frame(0)

        parcels = ''
        for i in range(5):
            parcel_id = self.v_driver.find_element_by_xpath(
                ".//*[@id='parentTablesContainer']/div[1]/table/tbody/tr[" + str(i + 1) + "]/td[2]/span").text
            parcels = parcels + parcel_id + "\n"
        print(parcels)
        file = open("..\Applications\Workflows\ProcessTestFiles\AppResources\parcelIDsforSearch.txt", "w+")
        file.write(parcels)

    def search_by_parcel_id(self,parcel_id):
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        print("in search_by_parcel_id method")
        self.v_driver.get("https://commerce.spscommerce.com/transaction-tracker/prod/transactions/")
        time.sleep(4)
        self.v_driver.switch_to.frame(0)
        # parcel_id_textbox = self.driver.find_element_by_xpath(
        #     "html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/label/div/div[2]/input")
        so.send_text_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/label/div/div[2]/input",parcel_id)

        # parcel_id_textbox.send_keys(parcel_id)
        time.sleep(4)
        # search_btn = self.driver.find_element_by_xpath(
        #     "html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[5]/div/button")
        so.click_element_by_xpath("html/body/div[1]/section/section/div/div/section/div/div[1]/div[2]/div[5]/div/button")

        # search_btn.click()
        time.sleep(4)
        first_parcel_id = self.v_driver.find_element_by_xpath(
            ".//*[@id='parentTablesContainer']/div[1]/table/tbody/tr/td[2]/span/a ")
        so.click_element_by_xpath(".//*[@id='parentTablesContainer']/div[1]/table/tbody/tr/td[2]/span/a")

        #first_parcel_id.click()
        time.sleep(3)

    def save_page_source_and_generate_parcel_id_from_dom(self):
        print("in save_page_source_and_generate_parcel_id_from_dom method")
        # print(page_source)
        file = open("D://TTcode.txt", "w+")
        # total_page_source=driver.page_source
        file.write(self.v_driver.page_source)
        file.close()
        # time.sleep(2)
        file = open("D://TTcode.txt", "r")
        # print(file.read())
        # match='<a href="" ng-click="vm.gotoAnchor('parcel-' + parcel.parcel_uid)" track-link-event="quick-navigation" title="Go to parcel'
        match = "ng-click"
        for line in file:
            if re.match("(.*)vm.gotoAnchor(.*)", line):
                # print(line)
                numbers = re.findall('\d+', line)
                # print(numbers)
                d = numbers
                d = str(d)
                for numbers in ["]"]:
                    if numbers in d:
                        d = d.replace(numbers, "")
                        e = d
                        e = str(e)
                        for numbers in ["["]:
                            if numbers in e:
                                e = e.replace(numbers, "")
                                f = e
                                f = str(f)
                                for numbers in ["'"]:
                                    if numbers in f:
                                        f = f.replace(numbers, "")
                                        return f

                break

    def generate_xpath_for_TT(self,generated_parcel_id):
        print("in generate_xpath_for_TT method")
        # time.sleep(2)
        xpath = ".//*[@id='parcel-" + str(generated_parcel_id) + "']/div/div/div[1]/div/div/a/i[1]"
        return xpath

    # This method performs the validation of map extensions
    def add_extensions(self, counter):

        # Get the versions of the supplier and retailer from Input sheet
        eo = ExcelOperations(self.v_task_type, self.v_data_sheet)

        supplier_map_version = self.v_data_sheet.cell(row=counter, column=9).value
        retailer_map_version = self.v_data_sheet.cell(row=counter, column=10).value

        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)

        version_check = msgbox("Supplier_Version"+supplier_map_version+"             "+"Retailer_Version"+retailer_map_version)


        # # print(self.v_driver.current_window_handle)
        # if  so.check_exists_by_xpath(V_configure_extensions_xpath):
        #     so.click_element_by_xpath(V_configure_extensions_xpath)
        #     time.sleep(1)
        #     self.v_driver.switch_to.window(self.v_driver.window_handles[-1])
        #     time.sleep(1)
        #     self.v_driver.switch_to.frame(0)
        #     self.lo.log_to_file("INFO", "Checking the extensions")



        self.v_driver.switch_to.window(self.v_driver.window_handles[0])
