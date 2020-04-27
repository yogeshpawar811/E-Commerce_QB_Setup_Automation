'''

@ Author - Aditya Datar
@ Creation date - 09/14/2018
@ Description - Main Script of Review Setup
'''

import openpyxl
from builtins import set
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.ExcelOperations import ExcelOperations
from Utilites.SeleniumOperations import SeleniumOperations
from selenium.common.exceptions import NoSuchElementException
from Applications.Workflows.ProcessTestFiles.AppResources import CommonLocators
from Applications.Workflows.ProcessTestFiles.Scripts.Process_Test_File_850 import Process_Test_file_850
from Applications.Workflows.ProcessTestFiles.Scripts import Process_Test_File_Download
from Utilites.TransactionTrackerOperations import TransactionTrackerOperations

import time
import re

class Process_Test_Files:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')
        # self.v_data_wb = openpyxl.load_workbook(AppConstants.PROCESS_TEST_FILES_INPUT_PATH)
        self.lo = lo
        self.v_username = username

    def execute_main(self):
        # self.lo.log_to_file("INFO", "Login in to DC4 Pre_Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type, self.v_driver, self.lo)
        self.v_data_sheet = self.v_input_wb.get_sheet_by_name('PROCESS_TEST_FILES_INPUT')

        total_rows = self.v_data_sheet.max_row
        self.v_driver.maximize_window()
        lg.login("Launchpad")
        time.sleep(3)
        self.v_driver.get("https://commerce.spscommerce.com/transaction-tracker/prod/transactions/")
        time.sleep(5)
        tto=TransactionTrackerOperations(self.v_task_type, self.v_driver, self.lo)

        # tto.search_process_for_process_test_file("Spring Silver Foods Inc","C&S Wholesale","850","10/02/2018")
        tto.search_by_parcel_id_and_download("10944250370")

        # Switch to other window
        # self.v_driver.execute_script("window.open('about:blank', 'tab2');")
        # self.v_driver.switch_to.window("tab2")
        #
        # lg.login("DC4 PreProd")

        #self.v_driver.maximize_window()






        # for row_count in range(2,total_rows+1):
        #
        #     v_supplier = self.v_data_sheet.cell(row=row_count, column=1).value
        #     v_retailer = self.v_data_sheet.cell(row=row_count, column=2).value
        #     v_doc_type = self.v_data_sheet.cell(row=row_count, column=3).value
        #     if v_doc_type == CommonLocators.PO_File:
        #         ptf_850 = Process_Test_file_850(self.v_task_type,self.lo,self.v_username,self.v_input_wb, self.v_driver)
        #         ptf_850.execute_main(v_supplier, v_retailer, v_doc_type,row_count)
        #
        #     else:
        #         self.lo.log_to_file("ERROR", "Document is not found")








