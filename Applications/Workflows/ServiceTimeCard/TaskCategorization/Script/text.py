


import openpyxl
from Applications.Workflows.ServiceTimeCard.TaskCategorization.AppResources import ElementLocators
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
from Applications.Workflows.ServiceTimeCard.TaskCategorization.Script.Salesforce import Salesforce
import time
import openpyxl
from Applications.Workflows.ServiceTimeCard.TaskCategorization.AppResources import ElementLocators
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
import time
from selenium.webdriver.common.keys import Keys



def get_maps(doc_type,supplier_version,retailer_version):
    v_input_wb = openpyxl.load_workbook("C:/Users/yogesh.pawar/Desktop/Conversion_Maps.xlsx")
    v_input_sheet = v_input_wb.get_sheet_by_name("Sheet1")

    for i in range(v_input_sheet.max_row):
        if str(v_input_sheet.cell(row=(i+1), column=1).value)==str(doc_type) and str(v_input_sheet.cell(row=(i+1), column=2).value)==str(supplier_version) and str(v_input_sheet.cell(row=(i+1), column=3).value)== str(retailer_version):
            iteration=str(v_input_sheet.cell(row=(i+1), column=4).value)
            maps=str(v_input_sheet.cell(row=(i+1), column=5).value)
    return iteration,maps


def add_extention(doc_type,supplier_version,retailer_version):
    so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)

    value = get_maps(doc_type,supplier_version,retailer_version)
    total_maps=value[0]
    all_maps=value[1]
    maps=all_maps.split(",")

    so.click_element_by_xpath(".//*[@id='form1:table1:commandButton1']/img")#click on add extention btn
    so.click_element_by_xpath(".//.[contains(text(),'1080')]//preceding::*[contains(@type,'radio')][1]")#1080 map extention
    so.click_element_by_xpath(".//*[@id='form1:table1:commandButton2_adfr_adfr']/img")#choose button

    for i in range(int(total_maps)):
        so.click_element_by_xpath(".//*[@id='form1:table1dd"+str(i)+"']/img")#start from 0 and execute 3 time

    for k in range(int(total_maps)):
        so.send_text_by_xpath(".//*[@id='form1:table1:"+str(k)+":table2:0:outputText22']",maps[k])#start from 0 and execute 3 time

    so.click_element_by_xpath(".//*[@id='form1:commandButton2_adfr']/img")#save button



