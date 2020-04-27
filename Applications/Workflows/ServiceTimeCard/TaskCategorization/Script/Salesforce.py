import openpyxl
from Applications.Workflows.ServiceTimeCard.TaskCategorization.AppResources import ElementLocators
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations
import time
from selenium.webdriver.common.keys import Keys


class Salesforce:

    def __init__(self, task_type, lo, username,v_Browser):


        self.v_input_wb = openpyxl.load_workbook(ElementLocators.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name("Input")
        self.log = lo
        self.v_username = username
        self.v_Browser=v_Browser
        self.so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)



    def login_to_sailpoint(self):
        print("login_to_sailpoint")
        so = SeleniumOperations(self.v_task_type, self.v_Browser, self.log)

        self.v_Browser.get(ElementLocators.SALESFORCE_URL)
        so.click_element_by_xpath(".//*[@id='idp_section_buttons']/button[2]")
        # so.click_element_by_xpath(".//*[@id='cancel_idp_hint']")#Log In with a Different Account link
        # so.click_element_by_xpath(".//*[contains(text(),'SailPoint')]")
        # so.click_element_by_xpath(ElementLocators.SAILPOINT_BTN)
        # self.v_Browser.get("https://iam.spscommerce.com/login/login?spEntityID=https%3A%2F%2Fspscommerce.my.salesforce.com&goto=https%3A%2F%2Fiam-sso.spscommerce.com%2Fsso%2FSSORedirect%2FmetaAlias%2Fspscommerce%2Fidp%3FReqID%3D_2CAAAAXCfPa15ME8wMGcwMDAwMDA0Qzk4AAAA3kV67IWmbs12h6m088CMAf07tJcrGOLzEuC6k0Aq2epzV0oaRieDf2U4LBX0Ve0lPuXoZm0y4nxzD4vTrqr7X4mue0_5qLojuNHUxFCj61_mB3PZQj97LDNHpjxkZunTSBRH90cSn5xydi1kFAyU8PVF4s8BKuzm6j8kJzse7O_CephJERN1e6PoPj0VKwaWkFnvGnr-YNZyL64Uy3nUyttHGaMkHw4OYrepEuTK31x1D7qp0MwTmiK2l4_ZqOdVBw%26index%3Dnull%26acsURL%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%253Fso%253D00D300000000bzv%26spEntityID%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%26binding%3Durn%253Aoasis%253Anames%253Atc%253ASAML%253A2.0%253Abindings%253AHTTP-POST")
        # time.sleep(2)
        # self.v_Browser.get("https://iam.spscommerce.com/login/login?spEntityID=https%3A%2F%2Fspscommerce.my.salesforce.com&goto=https%3A%2F%2Fiam-sso.spscommerce.com%2Fsso%2FSSORedirect%2FmetaAlias%2Fspscommerce%2Fidp%3FReqID%3D_2CAAAAXCfPa15ME8wMGcwMDAwMDA0Qzk4AAAA3kV67IWmbs12h6m088CMAf07tJcrGOLzEuC6k0Aq2epzV0oaRieDf2U4LBX0Ve0lPuXoZm0y4nxzD4vTrqr7X4mue0_5qLojuNHUxFCj61_mB3PZQj97LDNHpjxkZunTSBRH90cSn5xydi1kFAyU8PVF4s8BKuzm6j8kJzse7O_CephJERN1e6PoPj0VKwaWkFnvGnr-YNZyL64Uy3nUyttHGaMkHw4OYrepEuTK31x1D7qp0MwTmiK2l4_ZqOdVBw%26index%3Dnull%26acsURL%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%253Fso%253D00D300000000bzv%26spEntityID%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%26binding%3Durn%253Aoasis%253Anames%253Atc%253ASAML%253A2.0%253Abindings%253AHTTP-POST")
        # self.v_Browser.get("https://iam.spscommerce.com/login/login?spEntityID=https%3A%2F%2Fspscommerce.my.salesforce.com&goto=https%3A%2F%2Fiam-sso.spscommerce.com%2Fsso%2FSSORedirect%2FmetaAlias%2Fspscommerce%2Fidp%3FReqID%3D_2CAAAAXCfPa15ME8wMGcwMDAwMDA0Qzk4AAAA3kV67IWmbs12h6m088CMAf07tJcrGOLzEuC6k0Aq2epzV0oaRieDf2U4LBX0Ve0lPuXoZm0y4nxzD4vTrqr7X4mue0_5qLojuNHUxFCj61_mB3PZQj97LDNHpjxkZunTSBRH90cSn5xydi1kFAyU8PVF4s8BKuzm6j8kJzse7O_CephJERN1e6PoPj0VKwaWkFnvGnr-YNZyL64Uy3nUyttHGaMkHw4OYrepEuTK31x1D7qp0MwTmiK2l4_ZqOdVBw%26index%3Dnull%26acsURL%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%253Fso%253D00D300000000bzv%26spEntityID%3Dhttps%253A%252F%252Fspscommerce.my.salesforce.com%26binding%3Durn%253Aoasis%253Anames%253Atc%253ASAML%253A2.0%253Abindings%253AHTTP-POST")
        so.send_text_by_xpath(ElementLocators.SAILPOINT_USERNAME_TEXTBOX, ElementLocators.SAILPOINT_CREDENTIAL_USENAME)
        so.send_text_by_xpath(ElementLocators.SAILPOINT_PASSWORD_TEXTBOX, ElementLocators.SAILPOINT_CREDENTIAL_PASSWORD)
        so.click_element_by_xpath(ElementLocators.SAILPOINT_LOGIN_BTN)

    def open_FS_Team_Queue(self):
        print("open_FS_Team_Queue")
        time.sleep(5)
        self.v_Browser.get("https://spscommerce.my.salesforce.com/")
        time.sleep(5)
        self.v_Browser.get("https://spscommerce.my.salesforce.com/")
        self.v_Browser.get(ElementLocators.QUEUE_URL)
        time.sleep(7)

        if self.so.check_exists_by_xpath(ElementLocators.CROSS_SWITCH_BOX) == True:
            self.so.click_element_by_xpath(ElementLocators.CROSS_SWITCH_BOX)

        self.v_Browser.get(ElementLocators.QUEUE_URL)
        time.sleep(2)
        self.so.click_element_by_xpath(ElementLocators.QUEUE_LIST)
        self.so.click_element_by_xpath(ElementLocators.FS_TEAM_QUEUE)
        if self.so.check_exists_by_xpath(ElementLocators.GO_BTN)==True:
          self.so.click_element_by_xpath(ElementLocators.GO_BTN)

    def get_all_ticket_info(self):
        print("get_all_ticket_info")
        all_tickets = self.v_Browser.find_elements_by_xpath(ElementLocators.ALL_CASES)
        Case_number_and_salesforce_id_array=[]
        for ii in all_tickets:
            # get salceforce ID
            try:
                Salesforce_ID = ii.get_attribute("id").split("_")[0]
                Case_number = ii.text
                Case_number_and_salesforce_id_array.append(Case_number+"_"+Salesforce_ID)
                # self.open_case_in_new_tab(Salesforce_ID)
            except:
                print("error occured in getting Salesforce_ID")
        return Case_number_and_salesforce_id_array

    def return_caseno_and_salesforceno(self,i,Case_number_and_salesforce_id_array):
        Case_number = Case_number_and_salesforce_id_array[i].split("_")[0]
        salesforce_id = Case_number_and_salesforce_id_array[i].split("_")[1]
        return Case_number,salesforce_id


    def open_cases_and_get_info(self,Case_number_and_salesforce_id_array):

        print("open_cases_and_get_info")
        # j=0
        # print(Case_number_and_salesforce_id_array)
        for i in range(len(Case_number_and_salesforce_id_array)):
            print("i value")
            print(i)
            data=self.return_caseno_and_salesforceno(i,Case_number_and_salesforce_id_array)
            # print(Case_number_and_salesforce_id_array)
            Case_number=data[0]
            salesforce_id=data[1]
            print("==========================================")
            print(Case_number)
            print(salesforce_id)
            print("==========================================")
            # time.sleep(5)
            print(Case_number_and_salesforce_id_array)

            # salesforce_id_array_task=Case_number_and_salesforce_id_array[i]
            # print(i)
            # print(salesforce_id)
            self.v_Browser.get("https://spscommerce.my.salesforce.com/"+str(salesforce_id))
            time.sleep(1)
            subject=self.v_Browser.find_element_by_xpath(".//*[contains(@id,'Subject')]").get_attribute("value").lower()

            # subject=self.so.get_text_by_xpath(".//*[contains(@id,'Subject')]").lower()
            # print(subject)
            # if "document processing error" in subject:
            #     findings="Document Processing Error"
            # elif "document error" in subject:
            #     findings = "Document Processing Error for document error"
            # elif "processing error" in subject:
            #     findings = "Document Processing Error for processing error"
            # else:
            #     findings="Unknown Error"

            mail_count=len(self.v_Browser.find_elements_by_xpath(".//*[contains(@class,'feeditemaux cxfeeditemaux EmailMessageAuxBody')]"))

            all_mails = self.v_Browser.find_elements_by_xpath(".//*[contains(@class,'feeditemaux cxfeeditemaux EmailMessageAuxBody')]")
            # Case_number_and_salesforce_id_array = []
            mail1=" "
            mail2=" "
            mail3=" "
            mail4=" "
            mail5=" "
            mail6 = " "
            mail7 = " "
            mail8 = " "
            mail9 = " "
            mail10 = " "
            mail11 = " "
            mail12 = " "
            mail13 = " "
            mail14 = " "
            mail15 = " "
            mail16 = " "
            mail17 = " "
            mail18 = " "
            mail19 = " "
            mail20 = " "
            mail21 = " "
            mail22 = " "
            mail23 = " "
            mail24 = " "
            mail25 = " "

            k=0
            for ii in all_mails:
                k=k+1
                mail_text=ii.text
                if k==1:
                    mail1=mail_text
                elif k==2:
                    mail2=mail_text
                elif k==3:
                    mail3=mail_text
                elif k==4:
                    mail4=mail_text
                elif k==5:
                    mail5=mail_text
                elif k==6:
                    mail6=mail_text
                elif k==7:
                    mail7=mail_text
                elif k==8:
                    mail8=mail_text
                elif k==9:
                    mail9=mail_text
                elif k==10:
                    mail10=mail_text
                elif k==11:
                    mail11=mail_text
                elif k==12:
                    mail12=mail_text
                elif k==13:
                    mail13=mail_text
                elif k==14:
                    mail14=mail_text
                elif k==15:
                    mail15=mail_text
                elif k==16:
                    mail16=mail_text
                elif k==17:
                    mail17=mail_text
                elif k==18:
                    mail18=mail_text
                elif k==19:
                    mail19=mail_text
                elif k==20:
                    mail20=mail_text
                elif k==21:
                    mail21=mail_text
                elif k==22:
                    mail22=mail_text
                elif k==23:
                    mail23=mail_text
                elif k==24:
                    mail24=mail_text
                elif k==25:
                    mail25=mail_text

            self.save_findings_in_excel(i,Case_number,salesforce_id,"findings",subject,mail1,mail2,mail3,mail4,mail5)


    def save_findings_in_excel(self,i,Case_number,salesforce_id,findings,subject,mail1=None,mail2=None,mail3=None,mail4=None,mail5=None,mail6=None,mail7=None,mail8=None,mail9=None,mail10=None,mail11=None,mail12=None,mail13=None,mail14=None,mail15=None,mail16=None,mail17=None,mail18=None,mail19=None,mail20=None,mail21=None,mail22=None,mail23=None,mail24=None,mail25=None):
        row_value=i+2
        self.v_input_sheet.cell(row=row_value, column=1).value = Case_number
        self.v_input_sheet.cell(row=row_value, column=2).value = salesforce_id
        self.v_input_sheet.cell(row=row_value, column=3).value = findings
        self.v_input_sheet.cell(row=row_value, column=4).value = subject
        self.v_input_sheet.cell(row=row_value, column=5).value = mail1
        self.v_input_sheet.cell(row=row_value, column=6).value = mail2
        self.v_input_sheet.cell(row=row_value, column=7).value = mail3
        self.v_input_sheet.cell(row=row_value, column=8).value = mail4
        self.v_input_sheet.cell(row=row_value, column=9).value = mail5
        self.v_input_sheet.cell(row=row_value, column=10).value = mail6
        self.v_input_sheet.cell(row=row_value, column=9).value = mail7
        self.v_input_sheet.cell(row=row_value, column=9).value = mail8
        self.v_input_sheet.cell(row=row_value, column=9).value = mail9
        self.v_input_sheet.cell(row=row_value, column=9).value = mail10
        self.v_input_sheet.cell(row=row_value, column=9).value = mail11
        self.v_input_sheet.cell(row=row_value, column=9).value = mail12
        self.v_input_sheet.cell(row=row_value, column=9).value = mail13
        self.v_input_sheet.cell(row=row_value, column=9).value = mail14
        self.v_input_sheet.cell(row=row_value, column=9).value = mail15
        self.v_input_sheet.cell(row=row_value, column=9).value = mail16
        self.v_input_sheet.cell(row=row_value, column=9).value = mail17
        self.v_input_sheet.cell(row=row_value, column=9).value = mail18
        self.v_input_sheet.cell(row=row_value, column=9).value = mail19
        self.v_input_sheet.cell(row=row_value, column=9).value = mail20
        self.v_input_sheet.cell(row=row_value, column=9).value = mail21
        self.v_input_sheet.cell(row=row_value, column=9).value = mail22
        self.v_input_sheet.cell(row=row_value, column=9).value = mail23
        self.v_input_sheet.cell(row=row_value, column=9).value = mail24
        self.v_input_sheet.cell(row=row_value, column=9).value = mail25


        self.v_input_wb.save(ElementLocators.INPUT_FILE_PATH)





