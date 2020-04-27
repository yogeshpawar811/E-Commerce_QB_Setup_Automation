'''
@ Author - Yogesh Pawar
@ Creation date - 01/10/2020
@ Description - Main Test Script of Document Processing Error
'''

import time
import openpyxl
from selenium import webdriver
from Applications.Workflows.ServiceTimeCard.RPA.AppResources import ElementLocators
from Applications.Workflows.ServiceTimeCard.DocumentProcessing.Scripts.TransactionTrackerActions import TransactionTrackerActions



class Main_Test_Script_DocumentProcessing:
    def __init__(self, task_type, lo, username):
        self.v_input_wb = openpyxl.load_workbook(ElementLocators.INPUT_FILE_PATH)
        self.v_task_type = task_type
        self.v_input_sheet = self.v_input_wb.get_sheet_by_name("Input")
        self.log = lo
        self.v_username = username


    def execute_main(self):

        # chrome_option = Options()
        # chrome_option.add_argument("--user-data-dir=C:/Users/Yogesh.Pawar/AppData/Local/Google/Chrome/User Data")
        # self.v_Browser = webdriver.Chrome(executable_path=ElementLocators.BROWSER_DRIVER, chrome_options=chrome_option)


        self.v_Browser = webdriver.Chrome(ElementLocators.BROWSER_DRIVER)
        self.v_Browser.maximize_window()
        TransactionTrackerActions.Login_Launchpad(self,self.v_Browser)

        row_count = self.v_input_sheet.max_row
        time.sleep(5)
        TransactionTrackerActions.for_unpo(self,self.v_Browser, "696400")
        time.sleep(15)

        # for i in range(row_count-1):
        #     svc_no=self.v_input_sheet.cell(row=i+2,column=1).value
        #     row=i+2
        #     error_parcel_id=self.v_input_sheet.cell(row=row,column=2).value
        #
        #     self.v_Browser.get(TransactionTrackerActions.Generate_TT_Prod_URL(error_parcel_id))
        #     time.sleep(15)
        #
        #     self.v_Browser.switch_to.frame(0)
        #
        #
        #
        #     try:
        #         id=self.v_Browser.find_elements_by_xpath(".//*[@class='tile-document errored']")
        #         for ii in id:
        #             if "SIP" in ii.text:
        #                 print(ii.text)
        #                 print("errored SIP found")
        #                 self.v_Browser.find_element_by_xpath(".//*[@class='tile-document errored']/div[contains(.,'SIP')]").click()
        #                 print("clicked")
        #                 # time.sleep(6)
        #                 break
        #                 print(ii.text)
        #                 print("errored FEDS found")
        #                 self.v_Browser.find_elements_by_xpath(".//*[@class='tile-document errored']/div[contains(.,'FEDS')]").click()
        #
        #
        #                 # time.sleep(6)
        #                 break
        #             if "APP" in ii.text:
        #                 print(ii.text)
        #                 print("errored APP found")
        #
        #                 multiple_app = self.v_Browser.find_elements_by_xpath(".//*[@class='tile-document errored']/div[contains(.,'NetSuite')]")
        #
        #
        #                 # self.v_Browser.find_elements_by_xpath(".//*[@class='tile-document errored']/div[contains(.,'APP')]").click()
        #
        #                 # self.v_Browser.find_elements_by_xpath(".//*[@class='tile-document errored']//preceding::strong").click()
        #                 self.v_Browser.find_elements_by_xpath(".//*[@class='tile-document errored']/div[contains(.,'NetSuite')]").click()
        #                 # time.sleep(6)
        #                 break
        #             else:
        #                 print("Other error format found")
        #                 break
        #     except:
        #         print("NO ERROR FOUND!")
        #     time.sleep(6)
        #
        #     errors=TransactionTrackerActions.get_parcels_errors(self.v_Browser)
        #     TransactionTrackerActions.save_errors_in_excel(self,row,errors)
        #
        #
        #     time.sleep(5)
        #



        # self.v_Browser.find_element_by_xpath().text
