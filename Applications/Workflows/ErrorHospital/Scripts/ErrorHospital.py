'''

@ Author - Karan Pandya
@ Creation date - 08/29/2018
@ Description - Main Script of Error Hospital
'''
import time
import math
import datetime
import openpyxl
from Utilites.Login import Login
from Utilites import AppConstants
from selenium import webdriver
from Utilites.LogFileUtility import LogFileUtility
from Utilites.DC4_Utility import DC4_Utility
from Utilites.SeleniumOperations import SeleniumOperations
from Utilites.ReportFileUtility import ReportFileUtility
from Applications.Workflows.ErrorHospital.Scripts.ErrorHospital_Utility import ErrorHospital_Utility
from Utilites.ExcelOperations import ExcelOperations
from Applications.Workflows.ErrorHospital.AppResources import LocalElementLocator
from openpyxl import load_workbook
class ErrorHospital:
    def __init__(self, task_type, lo, username):
        self.v_task_type = task_type
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--incognito")
        self.v_driver = webdriver.Chrome(AppConstants.BROWSER_DRIVER, chrome_options= chrome_options)
        self.v_input_wb = openpyxl.load_workbook(AppConstants.INPUT_FILE_PATH)
        self.lo = lo
        self.v_username = username

    def execute_main(self):
        v_start_time = time.time()
        self.lo.log_to_file("INFO", "Login in to DC4 Prod")
        lg = Login(self.v_task_type, self.v_driver, self.v_input_wb, self.lo)
        so = SeleniumOperations(self.v_task_type,self.v_driver,self.lo)
        ehu =  ErrorHospital_Utility(self.v_task_type,self.v_driver,self.lo)
        rf = ReportFileUtility(self.v_task_type)
        output_sheet = self.v_input_wb.get_sheet_by_name('Output')
        input_sheet = self.v_input_wb.get_sheet_by_name('Input')
        eo = ExcelOperations(self.v_task_type, input_sheet)
        output_sheet_eo = ExcelOperations(self.v_task_type, output_sheet)
        output_sheet_curr_row = 2
        for index in range(2,input_sheet.max_row+1):
            list_duplicate_ticket_uid = []
            today = datetime.datetime.now()
            DD = datetime.timedelta(days=14)
            earlier = today - DD
            earlier_str = earlier.strftime("%Y-%m-%d")
            earlier_str = earlier_str+' 00:00:00'
            lg.login("DC4 Prod")
            v_error_title = eo.get_value(index, 1)
            if v_error_title == LocalElementLocator.ADHOC_ERROR_TITLE:
                v_ticket_uid = eo.get_value(index, 2)
                if v_ticket_uid is None:
                    v_ticket_uid = ' '
                ehu.adhoc_error_search(v_ticket_uid,earlier_str,' ')
            if so.check_exists_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH):
                so.click_element_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH)
            tuid_index = 2
            while(ehu.get_ticket_uid(tuid_index)):
                temp_ticket_uid = ehu.get_ticket_uid(tuid_index)
                temp_receiver = ehu.get_receiver_name(tuid_index)
                temp_sender = ehu.get_sender_name(tuid_index)
                if temp_ticket_uid not in list_duplicate_ticket_uid:
                    ehu.adhoc_error_search(temp_ticket_uid,earlier_str,' ')
                    v_tpid, v_doctype = ehu.get_info_from_description(tuid_index-2)
                    ehu.adhoc_error_search(' ', earlier_str,v_tpid)
                    if so.check_exists_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH):
                        so.click_element_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH)
                    temp_tuid_index = 2
                    while (ehu.get_ticket_uid(temp_tuid_index)):
                        list_duplicate_ticket_uid.append(ehu.get_ticket_uid(temp_tuid_index))
                        temp_tuid_index = temp_tuid_index+1
                    output_sheet_eo.set_value(output_sheet_curr_row, 1, temp_ticket_uid)
                    output_sheet_eo.set_value(output_sheet_curr_row, 2, temp_sender)
                    output_sheet_eo.set_value(output_sheet_curr_row, 3, temp_receiver)
                    output_sheet_eo.set_value(output_sheet_curr_row, 4, v_tpid)
                    output_sheet_eo.set_value(output_sheet_curr_row, 5, v_doctype)
                    output_sheet_curr_row = output_sheet_curr_row+1
                    self.v_input_wb.save(AppConstants.INPUT_FILE_PATH)
                    ehu.adhoc_error_search(v_ticket_uid, earlier_str, ' ')
                    if so.check_exists_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH):
                        so.click_element_by_xpath(LocalElementLocator.EH_SHOW_ALL_TAB_XPATH)
                tuid_index = tuid_index+1



        #so.click_element(AppConstants.DC4_TAB, "BY_NAME")
        #dc = DC4_Utility(self.v_task_type, self.v_driver, self.lo)
        #dc.company_search_by_name('Midlab Inc')
        #dc.company_search_by_ISA_ID('6166651648')
        #dc.company_search_by_TPID('620TSTWONDERTRE')

        self.v_driver.close()
        v_end_time = time.time()
        rf.update_sheet(self.v_username, 2, math.floor(v_end_time - v_start_time), str(datetime.date.today()))





