'''
@ Author - Karan Pandya
@ Creation date - 08/27/2018
@ Description - Login for all SPS tools
'''
from selenium import webdriver
from Utilites.SeleniumOperations import SeleniumOperations, AppConstants, LogFileUtility
import openpyxl
from Utilites.ExcelOperations import ExcelOperations
class Login:
    def __init__(self, v_task_type, v_driver, v_input_wb, lo):
        self.driver = v_driver
        self.task_type = v_task_type
        self.input_wb = v_input_wb
        self.lo = lo
        print("in init of Login")
    def login(self, v_tool):

        login_wb = openpyxl.load_workbook(AppConstants.LOGIN_ENVIRONMENT_LOCATOR_FILE_PATH)
        sheet_names = login_wb.get_sheet_names()
        input_credential_sheet = self.input_wb.get_sheet_by_name("Credentials")
        eo = ExcelOperations(self.task_type, input_credential_sheet)
        so = SeleniumOperations(self.task_type, self.driver, self.lo)
        if v_tool in sheet_names:
            login_sheet = login_wb.get_sheet_by_name(v_tool)
        else:
            self.lo.log_to_file("ERROR", "Error in Login.login(). Sheet name not found")
            return False
        tool_row  = eo.search_for_element_in_column(1,v_tool)
        username = input_credential_sheet.cell(row = tool_row, column = 2).value
        password = input_credential_sheet.cell(row = tool_row, column = 3).value
        username_locator = login_sheet.cell(row = 2, column = 2).value
        password_locator = login_sheet.cell(row=3, column=2).value
        button_locator = login_sheet.cell(row=4, column=2).value
        if v_tool == "DC4 Prod":
            Link = AppConstants.DC4_PROD_LINK
        elif v_tool == "DC4 PreProd":
            Link = AppConstants.DC4_PREPROD_LINK
        elif v_tool == "JIRA":
            Link = AppConstants.JIRA_LINK
        elif v_tool == "Launchpad":
            Link = AppConstants.LAUNCHPAD_LINK
        elif v_tool == "Salesforce":
            Link = AppConstants.SALESFORCE_LINK
        else:
            self.lo.log_to_file("ERROR", "Error in Login.login(). Invalid Link")
            return False
        self.driver.get(Link)
        if v_tool == "Launchpad":
            self.driver.switch_to.frame(0)

        if v_tool == "Salesforce":
            so.click_element_by_xpath(AppConstants.SAILPOINT_TAB)
            so.send_text_by_xpath(username_locator, username)
            so.send_text_by_xpath(password_locator, password)
            so.click_element_by_xpath(button_locator)
        else:
            so.send_text_by_name(username_locator, username)
            so.send_text_by_name(password_locator, password)
            so.click_element_by_xpath(button_locator)