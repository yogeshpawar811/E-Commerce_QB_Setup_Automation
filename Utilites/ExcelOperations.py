'''

@ Author - Karan Pandya
@ Creation date - 08/30/2018
@ Description - Methods for reading the input excel sheet.
'''

import openpyxl

class ExcelOperations :

    def __init__(self, task_type, input_sheet):
        self.task_type = task_type
        self.input_sheet = input_sheet
        self.max_row = self.input_sheet.max_row
        self.max_column = self.input_sheet.max_column

    def search_for_element_in_column(self, column_index, key):
        for i in range(1, self.max_row+1):
            temp = self.input_sheet.cell(row = i, column = column_index).value
            if temp == key:
                return i
        return 0

    def get_value(self, row_index, column_index):
        return self.input_sheet.cell(row = row_index, column = column_index).value

    def set_value(self, row_index, column_index, key):
        self.input_sheet.cell(row=row_index, column=column_index).value = key


